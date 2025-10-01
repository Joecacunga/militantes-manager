import re, json, os
import pandas as pd
from io import BytesIO

CAP_REGEX = re.compile(r'^CAP\d+$', re.IGNORECASE)
PHONE_REGEX = re.compile(r'^(?:\+244|244)?\s*0?(\d{9})$')

def normalizar_telefone(t):
    if not t or not str(t).strip():
        return ""
    s = str(t).strip()
    m = PHONE_REGEX.match(s.replace(" ",""))
    if m:
        return m.group(1)
    digits = re.sub(r'\D','',s)
    if len(digits) >= 9:
        return digits[-9:]
    return ""

def validar_cap(cap):
    if not cap or not str(cap).strip():
        return False
    s = str(cap).upper().replace(" ","").replace("-","")
    return bool(CAP_REGEX.match(s))

def carregar_base(path):
    if os.path.exists(path):
        try:
            with open(path,"r",encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def guardar_base(base, path):
    try:
        with open(path,"w",encoding="utf-8") as f:
            json.dump(base,f,ensure_ascii=False,indent=2)
        return True
    except Exception as e:
        print("Erro ao guardar base:",e)
        return False

def chave_cap_normalizada(cap):
    return str(cap).upper().replace(" ","").replace("-","")

def adicionar_militante(base, reg):
    cap_key = chave_cap_normalizada(reg.get("Nº CAP",""))
    if cap_key == "":
        return False, "Nº CAP inválido"
    if cap_key not in base:
        base[cap_key] = []
    tel = normalizar_telefone(reg.get("Telefone",""))
    for existing in base.get(cap_key,[]):
        if existing.get("Nome(s) Próprio(s)","").strip().upper() == reg.get("Nome(s) Próprio(s)","").strip().upper() \
           and existing.get("Último Nome","").strip().upper() == reg.get("Último Nome","").strip().upper() \
           and normalizar_telefone(existing.get("Telefone","")) == tel:
            return False, "Registo duplicado (Nome + Apelido + Telefone)"
    ordem = 1
    if base.get(cap_key):
        ordem = max([r.get("Nº Ordem",0) for r in base[cap_key]] or [0]) + 1
    reg_copy = reg.copy()
    reg_copy["Nº Ordem"] = ordem
    reg_copy["Telefone"] = tel
    base[cap_key].append(reg_copy)
    return True, "Adicionado"

def importar_excel_para_base(base, uploaded_file):
    stats = {"added":0, "skipped_invalid":0, "skipped_duplicates":0}
    try:
        if uploaded_file.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
    except Exception as e:
        raise RuntimeError(f"Falha a ler ficheiro: {e}")
    cols = {c.lower():c for c in df.columns}
    def get_col(*cands):
        for c in cands:
            if c in cols:
                return cols[c]
        return None
    col_nome = get_col("nome(s) próprio(s)","nome(s)","primeiro nome","nome","first name")
    col_apelido = get_col("último nome","apelido","last name","sobrenome")
    col_cap = get_col("nº cap","cap","num cap","n cap")
    col_tel = get_col("telefone","telemovel","phone","contact")
    col_cartao = get_col("cartao","cartão","has card","card")
    col_mun = get_col("municipio","município","municipio")
    col_com = get_col("comuna","commune")
    col_bairro = get_col("bairro","bairro")
    if not col_nome or not col_apelido or not col_cap or not col_tel:
        raise RuntimeError("Faltam colunas obrigatórias no ficheiro. Certifica-te que existe Nome(s), Último Nome, Nº CAP e Telefone.")
    for _, row in df.iterrows():
        try:
            reg = {
                "Nome(s) Próprio(s)": str(row.get(col_nome,"")).strip(),
                "Último Nome": str(row.get(col_apelido,"")).strip(),
                "Nº CAP": str(row.get(col_cap,"")).strip().upper().replace(" ","").replace("-",""),
                "Telefone": str(row.get(col_tel,"")).strip(),
                "Cartão": str(row.get(col_cartao,"")).strip() if col_cartao else "Não",
                "Província": str(row.get(col_mun,"")).strip() if col_mun else "",
                "Município": str(row.get(col_mun,"")).strip() if col_mun else "",
                "Comuna": str(row.get(col_com,"")).strip() if col_com else "",
                "Bairro": str(row.get(col_bairro,"")).strip() if col_bairro else ""
            }
            if not validar_cap(reg["Nº CAP"]):
                stats["skipped_invalid"] += 1
                continue
            tel_norm = normalizar_telefone(reg["Telefone"])
            if not tel_norm:
                stats["skipped_invalid"] += 1
                continue
            reg["Telefone"] = tel_norm
            ok, msg = adicionar_militante(base, reg)
            if ok:
                stats["added"] += 1
            else:
                stats["skipped_duplicates"] += 1
        except Exception:
            stats["skipped_invalid"] += 1
            continue
    return base, stats

def importar_paste_texto(base, text):
    stats = {"added":0, "skipped_invalid":0, "skipped_duplicates":0}
    lines = [l for l in text.splitlines() if l.strip()]
    header = None
    for i, line in enumerate(lines):
        if "\t" in line:
            parts = [p.strip() for p in line.split("\t")]
        elif "|" in line:
            parts = [p.strip() for p in line.split("|")]
        elif ";" in line:
            parts = [p.strip() for p in line.split(";")]
        else:
            parts = [p.strip() for p in line.split(",")]
        if i==0 and any(x.lower() in " ".join(parts).lower() for x in ["cap","telefone","nome"]):
            header = [p.lower() for p in parts]
            continue
        if header and len(parts)==len(header):
            d = dict(zip(header, parts))
            nome = d.get("nome(s) próprio(s)") or d.get("nome(s)") or d.get("nome") or d.get("first name") or ""
            apelido = d.get("último nome") or d.get("apelido") or d.get("last name") or ""
            cap = d.get("nº cap") or d.get("cap") or d.get("num cap") or ""
            tel = d.get("telefone") or d.get("phone") or ""
            cartao = d.get("cartao") or d.get("card") or "Não"
            mun = d.get("municipio") or ""
            com = d.get("comuna") or ""
            bairro = d.get("bairro") or ""
        else:
            partes = parts + [""]*8
            nome, apelido, cap, tel, cartao, mun, com, bairro = partes[:8]
        reg = {
            "Nome(s) Próprio(s)": nome.strip(),
            "Último Nome": apelido.strip(),
            "Nº CAP": str(cap).strip().upper().replace(" ","").replace("-",""),
            "Telefone": tel.strip(),
            "Cartão": cartao.strip() if cartao else "Não",
            "Província": mun.strip() if mun else "",
            "Município": mun.strip() if mun else "",
            "Comuna": com.strip() if com else "",
            "Bairro": bairro.strip() if bairro else ""
        }
        if not validar_cap(reg["Nº CAP"]):
            stats["skipped_invalid"] += 1
            continue
        teln = normalizar_telefone(reg["Telefone"])
        if not teln:
            stats["skipped_invalid"] += 1
            continue
        reg["Telefone"] = teln
        ok, msg = adicionar_militante(base, reg)
        if ok:
            stats["added"] += 1
        else:
            stats["skipped_duplicates"] += 1
    return base, stats

def gerar_relatorio_excel(base):
    from openpyxl import Workbook
    from io import BytesIO
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for cap, rows in base.items():
        ws = wb.create_sheet(title=str(cap)[:31])
        headers = ["Nº Ordem","Nome(s) Próprio(s)","Último Nome","Nº CAP","Telefone","Cartão","Província","Município","Comuna","Bairro"]
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h,"") for h in headers])
    ws = wb.create_sheet(title="RESUMO")
    total_militantes = sum(len(v) for v in base.values())
    total_caps = len(base.keys())
    com_cartao = sum(1 for lst in base.values() for r in lst if str(r.get("Cartão","")).lower()=="sim")
    sem_cartao = total_militantes - com_cartao
    ws.append(["Total Militantes", total_militantes])
    ws.append(["Total CAPs", total_caps])
    ws.append(["Com Cartão", com_cartao])
    ws.append(["Sem Cartão", sem_cartao])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def carregar_localidades(path):
    if os.path.exists(path):
        try:
            with open(path,"r",encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"Luanda":{}}
